# Excel táblából beolvasás, táblaforgatás (pivoting), csv fájl olvasás-írás

import openpyxl
import numpy as np

workbook = openpyxl.load_workbook(filename="numpy_7_test_1.xlsx")
ws = workbook.active
arr = np.empty((ws.max_row, ws.max_column), dtype=object)

print(arr)

# [[None None None None]
#  [None None None None]
#  [None None None None]
#  [None None None None]
#  [None None None None]]

# Ha előre tudjuk, mekkora tömbre lesz szükség, akkor célszerű azt előre lefoglalni.
# Ezzel sebességnövekedést érhetünk el.

for r, row in enumerate(ws.values):
    for c, value in enumerate(row):
        if value is None: value = ''
        arr[r, c] = str(value)

print(arr)

# [['' 'Budapest' 'New York' 'Hongkong']
#  ['2016' '90' '65' '102']
#  ['2017' '110' '80' '120']
#  ['2018' '200' '150' '190']
#  ['2019' '210' '170' '210']]

arr_2 = arr.transpose()
print(arr_2)

# [['' '2016' '2017' '2018' '2019']
#  ['Budapest' '90' '110' '200' '210']
#  ['New York' '65' '80' '150' '170']
#  ['Hongkong' '102' '120' '190' '210']]

# Ezt most elhelyezzük egy Excel táblázatban:

wb = openpyxl.Workbook()
ws = wb.active

for r, row in enumerate(arr_2):
    for c, value in enumerate(row):
        ws.cell(row = r + 1, column = c + 1).value = value

# és elmentjük:

wb.save("numpy_7_test_2.xlsx")

#############################

# Sokszor a leggyorsabb módja az adat exportálásának és importálásának csv fájl
# segítségével lehetséges.

# Excel vagy db ==> csv fájl ==> beolvasás, módosítás ==> csv fájl ==> Excel vagy db

import numpy as np
import csv

in_arr = []
csvfile = open('numpy_7_test_1.csv', newline='')
reader = csv.reader(csvfile, delimiter=';')
for row in reader:
    in_arr.append(row)
csvfile.close()

arr = np.array(in_arr)
print(arr)

# [['' 'Budapest' 'New York' 'Hongkong']
#  ['2016' '90' '65' '102']
#  ['2017' '110' '80' '120']
#  ['2018' '200' '150' '190']
#  ['2019' '210' '170' '210']]

arr_2 = arr.transpose()
print(arr_2)

# [['' '2016' '2017' '2018' '2019']
#  ['Budapest' '90' '110' '200' '210']
#  ['New York' '65' '80' '150' '170']
#  ['Hongkong' '102' '120' '190' '210']]

csvfile = open('numpy_7_test_2.csv', 'w', newline='')
writer = csv.writer(csvfile, delimiter='|')
for row in arr_2:
    writer.writerow(row)
csvfile.close()

#############################

